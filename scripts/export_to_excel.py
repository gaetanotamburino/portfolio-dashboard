"""
export_to_excel.py
───────────────────
Reads the latest snapshot from portfolio.db and writes it into the
'data' sheet of Portafogliov4.xlsm.

  Instruments table   row 1 (headers) + row 2 onwards   columns A:L
  Portfolio summary   2 rows below last instrument       columns A:I
"""

import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

BASE       = Path(__file__).resolve().parent.parent
DB_PATH    = BASE / "data" / "portfolio.db"
EXCEL_PATH = BASE / "Portafogliov4.xlsm"

PCT_FMT = "0.00%"
EUR_FMT = "#,##0.00"


# ── helpers ────────────────────────────────────────────────────────────────────

def _w(ws, row: int, col: int, value, fmt: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if fmt and value is not None:
        cell.number_format = fmt


def _pct(current, previous):
    """Safe percentage return between two prices."""
    if current is None or previous is None or previous == 0:
        return None
    return (current - previous) / previous


# ── data queries ───────────────────────────────────────────────────────────────

def _get_portfolio_summary(conn: sqlite3.Connection) -> pd.DataFrame:
    return pd.read_sql(
        """
        SELECT account, total_value, cost_basis, pnl_eur,
               pnl_pct_overall, ret_1d, ret_1w, ret_1m, ret_ytd
        FROM   portfolio_summary
        WHERE  date = (SELECT MAX(date) FROM portfolio_summary)
        ORDER  BY account
        """,
        conn,
    )


def _get_live_prices(conn: sqlite3.Connection) -> dict:
    """Return {yahoo_ticker: {price, market_state, fetched_at}} from live_prices."""
    try:
        rows = conn.execute(
            "SELECT symbol, price, market_state, fetched_at FROM live_prices"
        ).fetchall()
        return {r[0]: {"price": r[1], "market_state": r[2], "fetched_at": r[3]} for r in rows}
    except Exception:
        return {}


# ── Excel writer ───────────────────────────────────────────────────────────────

def _write_data_sheet(conn: sqlite3.Connection, wb) -> None:
    """
    Write to the 'data' sheet:
      - Instrument table  (headers row 1, data from row 2)  columns A:L
      - Portfolio summary (2 blank rows gap)                 columns A:I
    """
    ws = wb["data"] if "data" in wb.sheetnames else wb.create_sheet("data")

    latest = conn.execute("SELECT MAX(date) FROM positions").fetchone()[0]
    if not latest:
        return

    pos = pd.read_sql(
        """
        SELECT p.account, p.symbol, p.qty, p.avg_buy_price, p.close_price,
               i.name, i.broker, i.yahoo_ticker
        FROM   positions p
        JOIN   instruments i ON p.symbol = i.symbol
        WHERE  p.date = ?
        ORDER  BY p.account, p.symbol
        """,
        conn,
        params=[latest],
    )
    if pos.empty:
        return

    summary = _get_portfolio_summary(conn)

    # Compute per-ticker returns from price history
    tickers = pos["yahoo_ticker"].dropna().unique().tolist()
    ticker_stats: dict[str, dict] = {}
    if tickers:
        placeholders = ",".join("?" for _ in tickers)
        prices = pd.read_sql(
            f"""
            SELECT date, symbol, close
            FROM   prices
            WHERE  symbol IN ({placeholders})
            ORDER  BY symbol, date
            """,
            conn,
            params=tickers,
        )
        for ticker, grp in prices.groupby("symbol"):
            grp = grp.sort_values("date").reset_index(drop=True)
            cur  = grp.iloc[-1]["close"]
            n    = len(grp)
            year = int(latest[:4])
            prev_yr  = grp[grp["date"] <= f"{year - 1}-12-31"]
            ytd_base = prev_yr.iloc[-1]["close"] if not prev_yr.empty else None
            ticker_stats[ticker] = {
                # Fallback price (used when live is stale)
                "last_close":      cur,
                "last_close_date": grp.iloc[-1]["date"],
                # Denominators for live-price returns (indices shift by 1 vs
                # historical, because live is today and grp[-1] is yesterday)
                "ytd_base":  ytd_base,
                "denom_1w":  grp.iloc[-5]["close"]  if n >= 5  else None,
                "denom_1m":  grp.iloc[-21]["close"] if n >= 21 else None,
                # Pre-computed historical returns (used when live is stale)
                "ret_1d":  _pct(cur, grp.iloc[-2]["close"])  if n >= 2  else None,
                "ret_1w":  _pct(cur, grp.iloc[-6]["close"])  if n >= 6  else None,
                "ret_1m":  _pct(cur, grp.iloc[-22]["close"]) if n >= 22 else None,
                "ret_ytd": _pct(cur, ytd_base),
            }

    live = _get_live_prices(conn)

    # Clear a generous block to remove any stale data
    clear_rows = max(len(pos) + len(summary) + 20, 100)
    for r in range(1, clear_rows + 1):
        for c in range(1, 13):
            ws.cell(row=r, column=c).value = None

    # ── Instrument table ───────────────────────────────────────────────────────
    inst_headers = ["Symbol", "Instrument", "Account", "Broker", "Qty",
                    "Cost Basis", "Live Price",
                    "1D %", "1W %", "1M %", "YTD %", "Last Traded"]
    for c, h in enumerate(inst_headers, start=1):
        ws.cell(row=1, column=c, value=h)

    pipeline_ts = datetime.now()

    for i, (_, row) in enumerate(pos.iterrows()):
        r        = 2 + i
        yticker  = row["yahoo_ticker"]
        live_rec = live.get(yticker) if yticker else None
        stats    = ticker_stats.get(yticker, {})

        # Resolve price, timestamp, and period returns for col G / col L / col 8-11.
        # Fresh live price  → use it as numerator for all return windows.
        # Stale live price  → fall back to last historical close + pre-computed rets.
        price_val = None
        price_ts  = None
        ret_1d = stats.get("ret_1d")
        ret_1w = stats.get("ret_1w")
        ret_1m = stats.get("ret_1m")
        ret_ytd = stats.get("ret_ytd")

        if live_rec:
            try:
                fetched_dt = datetime.fromisoformat(live_rec["fetched_at"])
            except (TypeError, ValueError):
                fetched_dt = None

            if fetched_dt and fetched_dt.date() >= pipeline_ts.date():
                # Live price is from today — use it as the current price and
                # recompute returns: numerator = live, denominators shift by 1
                # (grp[-1] = yesterday, grp[-5] = 5 days ago, etc.)
                lp = live_rec["price"]
                price_val = lp
                price_ts  = live_rec["fetched_at"]
                ret_1d  = _pct(lp, stats.get("last_close"))
                ret_1w  = _pct(lp, stats.get("denom_1w"))
                ret_1m  = _pct(lp, stats.get("denom_1m"))
                ret_ytd = _pct(lp, stats.get("ytd_base"))
            else:
                # Stale — fall back to last historical close
                price_val  = stats.get("last_close")
                close_date = stats.get("last_close_date") or ""
                # Match fetched_at format ("%Y-%m-%dT%H:%M:%S", Europe/Rome).
                # US tickers (no exchange suffix) close at 22:00 Rome time;
                # European tickers close at 17:30.
                if yticker and "." not in yticker:
                    close_time = "22:00:00"
                else:
                    close_time = "17:30:00"
                price_ts = f"{close_date}T{close_time}" if close_date else None
        elif yticker is None:
            # No market ticker (synthetic positions, e.g. accrued bond interest) —
            # positions.close_price already holds the up-to-date computed value.
            price_val = row["close_price"]
            price_ts  = f"{latest}T00:00:00"

        _w(ws, r, 1,  row["symbol"])
        _w(ws, r, 2,  row["name"])
        _w(ws, r, 3,  row["account"])
        _w(ws, r, 4,  row["broker"])
        _w(ws, r, 5,  row["qty"])
        _w(ws, r, 6,  row["avg_buy_price"], EUR_FMT)
        _w(ws, r, 7,  price_val,            EUR_FMT)
        _w(ws, r, 8,  ret_1d,              PCT_FMT)
        _w(ws, r, 9,  ret_1w,              PCT_FMT)
        _w(ws, r, 10, ret_1m,              PCT_FMT)
        _w(ws, r, 11, ret_ytd,             PCT_FMT)
        _w(ws, r, 12, price_ts)

    print(f"[Step 4] Instruments: {len(pos)} row(s) written")

    # ── Portfolio summary (2 blank rows below last instrument) ─────────────────
    summary_start = 1 + len(pos) + 3   # row 1 header + len(pos) data + 2 blank + 1 for header

    summ_headers = ["Account", "Total Value", "Cost Basis", "P&L EUR",
                    "P&L %", "1D %", "1W %", "1M %", "YTD %"]
    for c, h in enumerate(summ_headers, start=1):
        ws.cell(row=summary_start, column=c, value=h)

    for i, (_, row) in enumerate(summary.iterrows()):
        r = summary_start + 1 + i
        _w(ws, r, 1, row["account"])
        _w(ws, r, 2, row["total_value"],     EUR_FMT)
        _w(ws, r, 3, row["cost_basis"],      EUR_FMT)
        _w(ws, r, 4, row["pnl_eur"],         EUR_FMT)
        _w(ws, r, 5, row["pnl_pct_overall"], PCT_FMT)
        _w(ws, r, 6, row["ret_1d"],          PCT_FMT)
        _w(ws, r, 7, row["ret_1w"],          PCT_FMT)
        _w(ws, r, 8, row["ret_1m"],          PCT_FMT)
        _w(ws, r, 9, row["ret_ytd"],         PCT_FMT)

    print(f"[Step 4] Portfolio summary: {len(summary)} account row(s) written")


def _write_charts_sheet(conn: sqlite3.Connection, wb) -> None:
    """
    Write to the 'charts' sheet:
      - Daily return time series (date x account ret_1d)  columns A:C
      - A line chart plotting each account's daily returns over time
    """
    ws = wb["charts"] if "charts" in wb.sheetnames else wb.create_sheet("charts")

    history = pd.read_sql(
        "SELECT date, account, ret_1d FROM portfolio_summary ORDER BY date",
        conn,
    )
    if history.empty:
        return

    pivot = (
        history.pivot_table(index="date", columns="account", values="ret_1d")
        .reset_index()
    )
    accounts = [c for c in pivot.columns if c != "date"]

    # Clear a generous block to remove any stale data/charts
    ws._charts = []
    clear_rows = max(len(pivot) + 10, 100)
    for r in range(1, clear_rows + 1):
        for c in range(1, len(accounts) + 3):
            ws.cell(row=r, column=c).value = None

    headers = ["Date"] + accounts
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    for i, (_, row) in enumerate(pivot.iterrows()):
        r = 2 + i
        _w(ws, r, 1, row["date"])
        for c, acct in enumerate(accounts, start=2):
            _w(ws, r, c, row[acct], PCT_FMT)

    last_row = 1 + len(pivot)

    chart = LineChart()
    chart.title = "Portfolio Daily Returns"
    chart.y_axis.title = "1D Return"
    chart.x_axis.title = "Date"
    chart.y_axis.numFmt = PCT_FMT
    chart.width = 28
    chart.height = 14

    data = Reference(ws, min_col=2, max_col=1 + len(accounts), min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, f"{chr(ord('A') + len(accounts) + 2)}2")

    print(f"[Step 4] Charts: {len(pivot)} daily return row(s) written for {', '.join(accounts)}")


def export(conn: sqlite3.Connection) -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, keep_vba=True)
    _write_data_sheet(conn, wb)
    _write_charts_sheet(conn, wb)

    try:
        wb.save(EXCEL_PATH)
        print(f"[Step 4] Saved -> {EXCEL_PATH.name}")
    except PermissionError:
        raise PermissionError(
            f"Cannot save — close '{EXCEL_PATH.name}' in Excel first, then re-run."
        )


def main() -> None:
    conn = sqlite3.connect(DB_PATH)
    try:
        export(conn)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
